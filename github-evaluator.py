import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import requests
import re
import os
import json
import time
import subprocess
import tempfile
import shutil
import csv
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, field
from datetime import datetime

# Import pour Excel (optionnel)
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

@dataclass
class ExecutionResult:
    """Résultat de l'exécution d'un programme"""
    success: bool
    output: str
    error: str
    execution_time: float
    exit_code: int

@dataclass
class CodeFileAnalysis:
    """Analyse d'un fichier de code"""
    filename: str
    language: str
    lines_of_code: int
    has_comments: bool
    has_functions: bool
    syntax_quality: float
    execution_result: Optional[ExecutionResult] = None
    compilation_success: bool = False
    note: float = 0.0

@dataclass
class DocFileAnalysis:
    """Analyse d'un fichier de documentation"""
    filename: str
    word_count: int
    line_count: int
    has_structure: bool
    has_code_blocks: bool
    technical_content: float
    note: float = 0.0

@dataclass
class StudentResult:
    """Résultat complet d'un étudiant"""
    group_number: int
    student_name: str
    matricule: str
    code_url: str
    doc_url: str
    note_code: float
    note_documentation: float
    note_execution: float
    note_finale: float
    code_files_count: int
    doc_files_count: int
    compilation_success: bool
    execution_success: bool

@dataclass
class GroupEvaluation:
    """Évaluation complète d'un groupe"""
    group_number: int
    code_repo_url: str
    doc_repo_url: str
    code_files: List[CodeFileAnalysis] = field(default_factory=list)
    doc_files: List[DocFileAnalysis] = field(default_factory=list)
    note_code: float = 0.0
    note_documentation: float = 0.0
    note_execution: float = 0.0
    note_finale: float = 0.0
    temps_analyse: float = 0.0

class CodeExecutor:
    """Gestionnaire d'exécution de code"""
    
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp(prefix="github_eval_")
        
        # Commandes de compilation par langage
        self.compile_commands = {
            'C': 'gcc {input} -o {output}',
            'C++': 'g++ {input} -o {output}',
            'Java': 'javac {input}',
        }
        
        # Commandes d'exécution par langage
        self.run_commands = {
            'C': '{executable}',
            'C++': '{executable}',
            'Python': 'python {input}',
            'Java': 'java {classname}',
            'JavaScript': 'node {input}',
        }
    
    def compile_c_cpp(self, source_path: str, language: str) -> Tuple[bool, str, str]:
        """Compile un fichier C/C++"""
        output_path = source_path.replace('.c', '').replace('.cpp', '') + '.exe'
        
        cmd = self.compile_commands[language].format(
            input=source_path,
            output=output_path
        )
        
        try:
            process = subprocess.run(
                cmd,
                shell=True,
                capture_output=True,
                text=True,
                timeout=30
            )
            
            success = process.returncode == 0
            return success, process.stdout, process.stderr
        
        except subprocess.TimeoutExpired:
            return False, "", "Compilation timeout (30s)"
        except Exception as e:
            return False, "", str(e)
    
    def execute_program(self, file_path: str, language: str) -> ExecutionResult:
        """Exécute un programme"""
        start_time = time.time()
        
        try:
            # Compilation si nécessaire
            if language in ['C', 'C++']:
                success, stdout, stderr = self.compile_c_cpp(file_path, language)
                if not success:
                    return ExecutionResult(
                        success=False,
                        output=stdout,
                        error=stderr,
                        execution_time=0,
                        exit_code=-1
                    )
                
                # Exécuter le binaire compilé
                executable = file_path.replace('.c', '').replace('.cpp', '') + '.exe'
                cmd = executable
            
            elif language == 'Python':
                cmd = f'python "{file_path}"'
            
            elif language == 'Java':
                # Compiler puis exécuter
                compile_result = subprocess.run(
                    f'javac "{file_path}"',
                    shell=True,
                    capture_output=True,
                    text=True,
                    timeout=30
                )
                if compile_result.returncode != 0:
                    return ExecutionResult(False, "", compile_result.stderr, 0, -1)
                
                classname = os.path.splitext(os.path.basename(file_path))[0]
                cmd = f'java -cp "{os.path.dirname(file_path)}" {classname}'
            
            elif language == 'JavaScript':
                cmd = f'node "{file_path}"'
            
            else:
                return ExecutionResult(False, "", "Langage non supporté pour exécution", 0, -1)
            
            # Exécuter avec timeout
            process = subprocess.run(
                cmd,
                shell=True,
                capture_output=True,
                text=True,
                timeout=40,
                input=""  # Pas d'entrée utilisateur
            )
            
            execution_time = time.time() - start_time
            
            return ExecutionResult(
                success=process.returncode == 0,
                output=process.stdout,
                error=process.stderr,
                execution_time=execution_time,
                exit_code=process.returncode
            )
        
        except subprocess.TimeoutExpired:
            return ExecutionResult(
                success=False,
                output="",
                error="Timeout d'exécution (10s)",
                execution_time=150.0,
                exit_code=-1
            )
        except Exception as e:
            return ExecutionResult(
                success=False,
                output="",
                error=str(e),
                execution_time=time.time() - start_time,
                exit_code=-1
            )
    
    def cleanup(self):
        """Nettoie les fichiers temporaires"""
        try:
            shutil.rmtree(self.temp_dir, ignore_errors=True)
        except:
            pass

class UniversalGitHubEvaluator:
    """Évaluateur avec exécution de code"""
    
    def __init__(self, log_callback=None):
        self.session = requests.Session()
        self.session.headers.update({
            'Accept': 'application/vnd.github.v3+json',
            'User-Agent': 'Universal-Academic-Evaluator'
        })
        self.log_callback = log_callback
        self.executor = CodeExecutor()
        
        self.code_extensions = {
            '.c': 'C', '.cpp': 'C++', '.cc': 'C++', '.cxx': 'C++',
            '.h': 'C Header', '.hpp': 'C++ Header', '.py': 'Python',
            '.java': 'Java', '.js': 'JavaScript', '.ts': 'TypeScript',
        }
        
        self.doc_extensions = {'.md', '.txt', '.rst', '.adoc'}
        
        self.technical_keywords = {
            'algorithme', 'algorithm', 'complexité', 'complexity',
            'fonction', 'function', 'variable', 'pointeur', 'pointer',
            'tableau', 'array', 'boucle', 'loop', 'structure', 'class'
        }
    
    def log(self, message):
        if self.log_callback:
            self.log_callback(message)
    
    def parse_github_url(self, url: str) -> Optional[Tuple[str, str, str]]:
        url = url.strip()
        pattern1 = r'https?://github\.com/([^/]+)/([^/]+)(?:/tree/[^/]+/(.+))?'
        match = re.match(pattern1, url)
        if match:
            owner, repo, path = match.groups()
            return owner, repo.replace('.git', ''), path or ''
        
        pattern2 = r'github\.com/([^/]+)/([^/]+)(?:/tree/[^/]+/(.+))?'
        match = re.match(pattern2, url)
        if match:
            owner, repo, path = match.groups()
            return owner, repo.replace('.git', ''), path or ''
        
        pattern3 = r'^([^/]+)/([^/]+)$'
        match = re.match(pattern3, url)
        if match:
            owner, repo = match.groups()
            return owner, repo.replace('.git', ''), ''
        
        return None
    
    def get_repo_contents(self, owner: str, repo: str, path: str = "") -> List[Dict]:
        url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"
        try:
            response = self.session.get(url, timeout=15)
            if response.status_code == 200:
                return response.json() if isinstance(response.json(), list) else []
            elif response.status_code == 403:
                self.log("⚠️ Limite API atteinte, pause...")
                time.sleep(60)
                return self.get_repo_contents(owner, repo, path)
            return []
        except:
            return []
    
    def get_file_content(self, owner: str, repo: str, file_path: str) -> str:
        url = f"https://api.github.com/repos/{owner}/{repo}/contents/{file_path}"
        try:
            response = self.session.get(url, timeout=15)
            if response.status_code == 200:
                file_data = response.json()
                if file_data.get('encoding') == 'base64':
                    import base64
                    return base64.b64decode(file_data['content']).decode('utf-8', errors='ignore')
        except:
            pass
        return ""
    
    def get_all_files_recursive(self, owner: str, repo: str, path: str = "") -> List[Dict]:
        all_files = []
        contents = self.get_repo_contents(owner, repo, path)
        for item in contents:
            if item['type'] == 'file':
                all_files.append(item)
            elif item['type'] == 'dir':
                sub_files = self.get_all_files_recursive(owner, repo, item['path'])
                all_files.extend(sub_files)
        return all_files
    
    def analyze_code_file(self, owner: str, repo: str, file_info: Dict) -> CodeFileAnalysis:
        filename = file_info['name']
        file_path = file_info['path']
        ext = os.path.splitext(filename)[1].lower()
        language = self.code_extensions.get(ext, 'Unknown')
        
        self.log(f"    📄 {filename} ({language})")
        
        content = self.get_file_content(owner, repo, file_path)
        
        if not content.strip():
            return CodeFileAnalysis(filename, language, 0, False, False, 0.0, None, False, 0.0)
        
        # Analyse statique
        lines = content.split('\n')
        lines_of_code = len([l for l in lines if l.strip() and not l.strip().startswith('//')])
        has_comments = '//' in content or '/*' in content or '#' in content
        has_functions = bool(re.search(r'def\s+\w+|function\s+\w+|\w+\s*\([^)]*\)\s*{|int\s+main', content))
        
        quality_score = 0.0
        indented_lines = sum(1 for l in lines if l.startswith('    ') or l.startswith('\t'))
        if lines_of_code > 0 and indented_lines / lines_of_code > 0.2:
            quality_score += 0.3
        if has_comments:
            quality_score += 0.3
        if has_functions:
            quality_score += 0.4
        
        # Note statique de base
        note = 0.0
        if lines_of_code > 0:
            note += 4.0
        if lines_of_code > 20:
            note += 2.0
        if has_comments:
            note += 3.0
        if has_functions:
            note += 3.0
        note += quality_score * 3.0
        
        # Exécution du code
        execution_result = None
        compilation_success = False
        
        if language in ['C', 'C++', 'Python', 'Java', 'JavaScript']:
            self.log(f"       🔧 Compilation/Exécution...")
            
            # Sauvegarder le fichier temporairement
            temp_file = os.path.join(self.executor.temp_dir, filename)
            with open(temp_file, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # Exécuter
            execution_result = self.executor.execute_program(temp_file, language)
            
            if execution_result.success:
                compilation_success = True
                note += 5.0  # Bonus pour exécution réussie
                self.log(f"       ✅ Exécution réussie ({execution_result.execution_time:.2f}s)")
            else:
                self.log(f"       ❌ Erreur: {execution_result.error[:100]}")
                note += 1.0  # Petit bonus pour tentative
        
        return CodeFileAnalysis(
            filename, language, lines_of_code, has_comments,
            has_functions, quality_score, execution_result,
            compilation_success, min(note, 20.0)
        )
    
    def analyze_doc_file(self, owner: str, repo: str, file_info: Dict) -> DocFileAnalysis:
        filename = file_info['name']
        file_path = file_info['path']
        
        self.log(f"    📖 {filename}")
        
        content = self.get_file_content(owner, repo, file_path)
        
        if not content.strip():
            return DocFileAnalysis(filename, 0, 0, False, False, 0.0, 0.0)
        
        word_count = len(content.split())
        line_count = len(content.split('\n'))
        has_structure = bool(re.search(r'^#+\s|^##|^===', content, re.MULTILINE))
        has_code_blocks = bool(re.search(r'```|`[^`]+`', content))
        
        content_lower = content.lower()
        tech_words = sum(1 for keyword in self.technical_keywords if keyword in content_lower)
        technical_content = min(tech_words / 10.0, 1.0)
        
        note = 0.0
        if word_count >= 50:
            note += 3.0
        if word_count >= 150:
            note += 2.0
        if word_count >= 300:
            note += 2.0
        if has_structure:
            note += 4.0
        if has_code_blocks:
            note += 3.0
        note += technical_content * 6.0
        
        return DocFileAnalysis(filename, word_count, line_count, has_structure,
                              has_code_blocks, technical_content, min(note, 20.0))
    
    def evaluate_code_repository(self, code_url: str) -> Tuple[List[CodeFileAnalysis], float, float]:
        self.log("\n🔧 Analyse du code...")
        parsed = self.parse_github_url(code_url)
        if not parsed:
            self.log(f"❌ URL invalide: {code_url}")
            return [], 0.0, 0.0
        
        owner, repo, path = parsed
        self.log(f"  📂 {owner}/{repo}")
        
        all_files = self.get_all_files_recursive(owner, repo, path)
        code_files = [f for f in all_files 
                     if os.path.splitext(f['name'])[1].lower() in self.code_extensions]
        
        self.log(f"  ✅ {len(code_files)} fichier(s) de code")
        
        analyses = []
        total_execution_score = 0
        execution_count = 0
        
        for file_info in code_files:
            analysis = self.analyze_code_file(owner, repo, file_info)
            analyses.append(analysis)
            self.log(f"       Note: {analysis.note:.1f}/20")
            
            if analysis.execution_result:
                execution_count += 1
                if analysis.execution_result.success:
                    total_execution_score += 20.0
                else:
                    total_execution_score += 5.0  # Points pour tentative
            
            time.sleep(0.1)
        
        note_code = sum(a.note for a in analyses) / len(analyses) if analyses else 0.0
        note_execution = total_execution_score / execution_count if execution_count > 0 else 0.0
        
        return analyses, note_code, note_execution
    
    def evaluate_doc_repository(self, doc_url: str) -> Tuple[List[DocFileAnalysis], float]:
        self.log("\n📚 Analyse de la documentation...")
        parsed = self.parse_github_url(doc_url)
        if not parsed:
            self.log(f"❌ URL invalide: {doc_url}")
            return [], 0.0
        
        owner, repo, path = parsed
        self.log(f"  📂 {owner}/{repo}")
        
        all_files = self.get_all_files_recursive(owner, repo, path)
        doc_files = [f for f in all_files 
                    if os.path.splitext(f['name'])[1].lower() in self.doc_extensions]
        
        self.log(f"  ✅ {len(doc_files)} fichier(s) de documentation")
        
        analyses = []
        for file_info in doc_files:
            analysis = self.analyze_doc_file(owner, repo, file_info)
            analyses.append(analysis)
            self.log(f"       Note: {analysis.note:.1f}/20")
            time.sleep(0.1)
        
        note_moyenne = sum(a.note for a in analyses) / len(analyses) if analyses else 0.0
        return analyses, note_moyenne
    
    def evaluate_group(self, group_number: int, code_url: str, doc_url: str) -> GroupEvaluation:
        self.log(f"\n{'='*60}")
        self.log(f"🏢 GROUPE {group_number}")
        self.log(f"{'='*60}")
        
        start_time = time.time()
        
        code_analyses, note_code, note_execution = self.evaluate_code_repository(code_url)
        doc_analyses, note_doc = self.evaluate_doc_repository(doc_url)
        
        # Note finale : 40% code statique + 30% exécution + 30% documentation
        note_finale = (note_code * 0.4) + (note_execution * 0.3) + (note_doc * 0.3)
        
        temps_analyse = time.time() - start_time
        
        self.log(f"\n📊 RÉSULTATS:")
        self.log(f"  Code (statique):   {note_code:.2f}/20")
        self.log(f"  Exécution:         {note_execution:.2f}/20")
        self.log(f"  Documentation:     {note_doc:.2f}/20")
        self.log(f"  NOTE FINALE:       {note_finale:.2f}/20")
        
        return GroupEvaluation(
            group_number, code_url, doc_url, code_analyses,
            doc_analyses, note_code, note_doc, note_execution, note_finale, temps_analyse
        )
    
    def cleanup(self):
        self.executor.cleanup()

class ExportManager:
    """Gestionnaire d'export CSV/Excel"""
    
    @staticmethod
    def export_to_csv(results: List[StudentResult], filename: str):
        """Exporte les résultats en CSV"""
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # En-têtes
            writer.writerow([
                'Groupe',
                'Nom Étudiant',
                'Matricule',
                'Note Code',
                'Note Exécution',
                'Note Documentation',
                'NOTE FINALE',
                'Fichiers Code',
                'Fichiers Doc',
                'Compilation OK',
                'Exécution OK',
                'URL Code',
                'URL Documentation'
            ])
            
            # Données
            for result in results:
                writer.writerow([
                    result.group_number,
                    result.student_name,
                    result.matricule,
                    f"{result.note_code:.2f}",
                    f"{result.note_execution:.2f}",
                    f"{result.note_documentation:.2f}",
                    f"{result.note_finale:.2f}",
                    result.code_files_count,
                    result.doc_files_count,
                    'Oui' if result.compilation_success else 'Non',
                    'Oui' if result.execution_success else 'Non',
                    result.code_url,
                    result.doc_url
                ])
    
    @staticmethod
    def export_to_excel(results: List[StudentResult], filename: str):
        """Exporte les résultats en Excel avec formatage"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl n'est pas installé")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résultats Évaluation"
        
        # Styles
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = [
            'Groupe', 'Nom Étudiant', 'Matricule', 'Note Code', 'Note Exécution',
            'Note Documentation', 'NOTE FINALE', 'Fichiers Code', 'Fichiers Doc',
            'Compilation', 'Exécution', 'URL Code', 'URL Documentation'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Données
        for row_idx, result in enumerate(results, 2):
            data = [
                result.group_number,
                result.student_name,
                result.matricule,
                result.note_code,
                result.note_execution,
                result.note_documentation,
                result.note_finale,
                result.code_files_count,
                result.doc_files_count,
                'Oui' if result.compilation_success else 'Non',
                'Oui' if result.execution_success else 'Non',
                result.code_url,
                result.doc_url
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = center_align
                cell.border = border
                
                # Colorer la note finale
                if col == 7:  # Note finale
                    if value >= 16:
                        cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif value >= 10:
                        cell.fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
        
        # Ajuster largeurs
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['L'].width = 40
        ws.column_dimensions['M'].width = 40
        
        # Statistiques en bas
        last_row = len(results) + 3
        ws.cell(last_row, 1, "STATISTIQUES").font = Font(bold=True, size=12)
        ws.cell(last_row + 1, 1, "Moyenne générale:")
        avg = sum(r.note_finale for r in results) / len(results)
        ws.cell(last_row + 1, 2, f"{avg:.2f}/20")
        
        ws.cell(last_row + 2, 1, "Note maximale:")
        ws.cell(last_row + 2, 2, f"{max(r.note_finale for r in results):.2f}/20")
        
        ws.cell(last_row + 3, 1, "Note minimale:")
        ws.cell(last_row + 3, 2, f"{min(r.note_finale for r in results):.2f}/20")
        
        wb.save(filename)

class MinimalistGitHubEvaluatorGUI:
    """Interface minimaliste et professionnelle"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("GitHub Academic Evaluator")
        self.root.geometry("1200x750")
        self.root.minsize(1000, 650)
        
        # Couleurs professionnelles
        self.colors = {
            'bg': '#f8fafc',
            'surface': '#ffffff',
            'primary': '#0f172a',
            'secondary': '#64748b',
            'accent': '#3b82f6',
            'success': '#22c55e',
            'warning': '#f59e0b',
            'danger': '#ef4444',
            'border': '#e2e8f0'
        }
        
        self.root.configure(bg=self.colors['bg'])
        
        # Données
        self.groups = []
        self.evaluator = None
        self.is_evaluating = False
        self.all_results = []
        self.log_queue = queue.Queue()
        
        self.setup_styles()
        self.create_gui()
        self.check_log_queue()
    
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        
        # Boutons
        style.configure('Accent.TButton',
                       background=self.colors['accent'],
                       foreground='white',
                       borderwidth=0,
                       padding=(16, 10),
                       font=('Inter', 10))
        
        style.map('Accent.TButton',
                 background=[('active', '#2563eb')])
        
        style.configure('Success.TButton',
                       background=self.colors['success'],
                       foreground='white',
                       borderwidth=0,
                       padding=(12, 8),
                       font=('Inter', 9))
        
        style.configure('Outline.TButton',
                       background=self.colors['surface'],
                       foreground=self.colors['primary'],
                       borderwidth=1,
                       padding=(12, 8),
                       font=('Inter', 9))
    
    def create_gui(self):
        # Header minimaliste
        header = tk.Frame(self.root, bg=self.colors['surface'], height=70)
        header.pack(fill='x', padx=0, pady=0)
        header.pack_propagate(False)
        
        title_frame = tk.Frame(header, bg=self.colors['surface'])
        title_frame.pack(side='left', padx=30, pady=20)
        
        tk.Label(title_frame,
                text="GitHub Academic Evaluator",
                bg=self.colors['surface'],
                fg=self.colors['primary'],
                font=('Inter', 16, 'bold')).pack(anchor='w')
        
        tk.Label(title_frame,
                text="Code execution • Documentation analysis • Automated grading",
                bg=self.colors['surface'],
                fg=self.colors['secondary'],
                font=('Inter', 9)).pack(anchor='w')
        
        # Bouton export en haut à droite
        export_frame = tk.Frame(header, bg=self.colors['surface'])
        export_frame.pack(side='right', padx=30)
        
        self.export_btn = ttk.Button(export_frame,
                                     text="📊 Export Results",
                                     style='Success.TButton',
                                     command=self.export_results,
                                     state='disabled')
        self.export_btn.pack()
        
        # Séparateur
        tk.Frame(self.root, bg=self.colors['border'], height=1).pack(fill='x')
        
        # Contenu principal
        main = tk.Frame(self.root, bg=self.colors['bg'])
        main.pack(fill='both', expand=True, padx=30, pady=20)
        
        # Gauche : Formulaire
        left = tk.Frame(main, bg=self.colors['bg'])
        left.pack(side='left', fill='both', expand=True, padx=(0, 15))
        
        self.create_form(left)
        
        # Droite : Console
        right = tk.Frame(main, bg=self.colors['bg'])
        right.pack(side='right', fill='both', expand=True, padx=(15, 0))
        
        self.create_console(right)
        
        # Footer minimaliste
        footer = tk.Frame(self.root, bg=self.colors['surface'], height=50)
        footer.pack(fill='x', side='bottom')
        footer.pack_propagate(False)
        
        tk.Frame(self.root, bg=self.colors['border'], height=1).pack(fill='x', side='bottom')
        
        self.status_label = tk.Label(footer,
                                     text="● Ready",
                                     bg=self.colors['surface'],
                                     fg=self.colors['success'],
                                     font=('Inter', 10))
        self.status_label.pack(side='left', padx=30)
        
        self.count_label = tk.Label(footer,
                                    text="0 groups",
                                    bg=self.colors['surface'],
                                    fg=self.colors['secondary'],
                                    font=('Inter', 10))
        self.count_label.pack(side='right', padx=30)
    
    def create_form(self, parent):
        # Card pour le formulaire
        card = tk.Frame(parent, bg=self.colors['surface'], relief='flat', bd=0)
        card.pack(fill='both', expand=True)
        
        # Padding intérieur
        content = tk.Frame(card, bg=self.colors['surface'])
        content.pack(fill='both', expand=True, padx=25, pady=25)
        
        # Titre section
        tk.Label(content,
                text="Add Groups",
                bg=self.colors['surface'],
                fg=self.colors['primary'],
                font=('Inter', 13, 'bold')).pack(anchor='w', pady=(0, 20))
        
        # Champ code
        tk.Label(content,
                text="Code Repository URL",
                bg=self.colors['surface'],
                fg=self.colors['secondary'],
                font=('Inter', 9)).pack(anchor='w', pady=(0, 5))
        
        self.code_entry = tk.Entry(content,
                                   font=('Consolas', 10),
                                   relief='solid',
                                   borderwidth=1,
                                   bg=self.colors['surface'],
                                   fg=self.colors['primary'],
                                   insertbackground=self.colors['accent'])
        self.code_entry.pack(fill='x', ipady=10, pady=(0, 15))
        
        # Champ doc
        tk.Label(content,
                text="Documentation Repository URL",
                bg=self.colors['surface'],
                fg=self.colors['secondary'],
                font=('Inter', 9)).pack(anchor='w', pady=(0, 5))
        
        self.doc_entry = tk.Entry(content,
                                  font=('Consolas', 10),
                                  relief='solid',
                                  borderwidth=1,
                                  bg=self.colors['surface'],
                                  fg=self.colors['primary'],
                                  insertbackground=self.colors['accent'])
        self.doc_entry.pack(fill='x', ipady=10, pady=(0, 20))
        
        # Boutons
        btn_frame = tk.Frame(content, bg=self.colors['surface'])
        btn_frame.pack(fill='x', pady=(0, 20))
        
        ttk.Button(btn_frame,
                  text="+ Add Group",
                  style='Accent.TButton',
                  command=self.add_group).pack(side='left', expand=True, fill='x', padx=(0, 5))
        
        ttk.Button(btn_frame,
                  text="Clear",
                  style='Outline.TButton',
                  command=self.clear_inputs).pack(side='right', expand=True, fill='x', padx=(5, 0))
        
        # Liste des groupes
        tk.Label(content,
                text="Groups Queue",
                bg=self.colors['surface'],
                fg=self.colors['primary'],
                font=('Inter', 11, 'bold')).pack(anchor='w', pady=(10, 10))
        
        list_frame = tk.Frame(content, bg=self.colors['surface'])
        list_frame.pack(fill='both', expand=True)
        
        scrollbar = tk.Scrollbar(list_frame, bg=self.colors['surface'])
        scrollbar.pack(side='right', fill='y')
        
        self.group_listbox = tk.Listbox(list_frame,
                                        font=('Consolas', 9),
                                        yscrollcommand=scrollbar.set,
                                        relief='solid',
                                        borderwidth=1,
                                        bg=self.colors['bg'],
                                        fg=self.colors['primary'],
                                        selectmode='single',
                                        highlightthickness=0)
        self.group_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.group_listbox.yview)
        
        # Boutons d'action
        action_frame = tk.Frame(content, bg=self.colors['surface'])
        action_frame.pack(fill='x', pady=(15, 0))
        
        ttk.Button(action_frame,
                  text="Remove Selected",
                  style='Outline.TButton',
                  command=self.remove_selected).pack(side='left', padx=(0, 5))
        
        self.eval_btn = ttk.Button(action_frame,
                                   text="▶ Start Evaluation",
                                   style='Success.TButton',
                                   command=self.start_evaluation)
        self.eval_btn.pack(side='right', expand=True, fill='x', padx=(5, 0))
    
    def create_console(self, parent):
        card = tk.Frame(parent, bg=self.colors['surface'])
        card.pack(fill='both', expand=True)
        
        content = tk.Frame(card, bg=self.colors['surface'])
        content.pack(fill='both', expand=True, padx=25, pady=25)
        
        tk.Label(content,
                text="Evaluation Console",
                bg=self.colors['surface'],
                fg=self.colors['primary'],
                font=('Inter', 13, 'bold')).pack(anchor='w', pady=(0, 15))
        
        console_frame = tk.Frame(content, bg='#0f172a')
        console_frame.pack(fill='both', expand=True)
        
        self.log_text = scrolledtext.ScrolledText(console_frame,
                                                   font=('JetBrains Mono', 9),
                                                   bg='#0f172a',
                                                   fg='#e2e8f0',
                                                   insertbackground='white',
                                                   relief='flat',
                                                   wrap='word',
                                                   state='disabled',
                                                   padx=15,
                                                   pady=15)
        self.log_text.pack(fill='both', expand=True)
        
        self.log_text.tag_config('success', foreground='#22c55e')
        self.log_text.tag_config('error', foreground='#ef4444')
        self.log_text.tag_config('warning', foreground='#f59e0b')
        self.log_text.tag_config('info', foreground='#3b82f6')
        self.log_text.tag_config('header', foreground='#a855f7', font=('JetBrains Mono', 9, 'bold'))
        
        btn_frame = tk.Frame(content, bg=self.colors['surface'])
        btn_frame.pack(fill='x', pady=(15, 0))
        
        ttk.Button(btn_frame,
                  text="Clear Console",
                  style='Outline.TButton',
                  command=self.clear_logs).pack(side='left')
        
        ttk.Button(btn_frame,
                  text="Save Logs",
                  style='Outline.TButton',
                  command=self.save_logs).pack(side='left', padx=(10, 0))
    
    def add_group(self):
        code = self.code_entry.get().strip()
        doc = self.doc_entry.get().strip()
        
        if not code or not doc:
            messagebox.showwarning("Missing Fields", "Please fill both URL fields")
            return
        
        self.groups.append((code, doc))
        self.group_listbox.insert('end', f"Group {len(self.groups)}: {code[:50]}...")
        
        self.clear_inputs()
        self.count_label.config(text=f"{len(self.groups)} groups")
        
        self.log_message(f"✅ Group {len(self.groups)} added", 'success')
    
    def remove_selected(self):
        sel = self.group_listbox.curselection()
        if not sel:
            return
        
        idx = sel[0]
        self.group_listbox.delete(idx)
        self.groups.pop(idx)
        self.count_label.config(text=f"{len(self.groups)} groups")
        
        self.log_message(f"🗑️ Group {idx + 1} removed", 'warning')
    
    def clear_inputs(self):
        self.code_entry.delete(0, 'end')
        self.doc_entry.delete(0, 'end')
        self.code_entry.focus()
    
    def log_message(self, message, tag='info'):
        self.log_text.config(state='normal')
        self.log_text.insert('end', message + '\n', tag)
        self.log_text.see('end')
        self.log_text.config(state='disabled')
    
    def clear_logs(self):
        self.log_text.config(state='normal')
        self.log_text.delete('1.0', 'end')
        self.log_text.config(state='disabled')
    
    def save_logs(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt")],
            initialfile=f"evaluation_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(self.log_text.get('1.0', 'end'))
            messagebox.showinfo("Success", f"Logs saved to:\n{filename}")
    
    def check_log_queue(self):
        try:
            while True:
                message = self.log_queue.get_nowait()
                self.log_message(message, self.determine_tag(message))
        except queue.Empty:
            pass
        
        self.root.after(100, self.check_log_queue)
    
    def determine_tag(self, message):
        if '✅' in message or 'success' in message.lower():
            return 'success'
        elif '❌' in message or 'erreur' in message.lower() or 'error' in message.lower():
            return 'error'
        elif '⚠️' in message or 'warning' in message.lower():
            return 'warning'
        elif '=' in message or '🏢' in message:
            return 'header'
        return 'info'
    
    def queue_log(self, message):
        self.log_queue.put(message)
    
    def start_evaluation(self):
        if not self.groups:
            messagebox.showwarning("No Groups", "Please add at least one group")
            return
        
        if self.is_evaluating:
            return
        
        response = messagebox.askyesno(
            "Start Evaluation",
            f"Evaluate {len(self.groups)} group(s)?\n\n"
            "This will:\n"
            "• Analyze code quality\n"
            "• Compile and execute programs\n"
            "• Evaluate documentation\n\n"
            "This may take several minutes."
        )
        
        if not response:
            return
        
        self.is_evaluating = True
        self.eval_btn.config(state='disabled')
        self.status_label.config(text="● Evaluating...", fg=self.colors['warning'])
        
        thread = threading.Thread(target=self.run_evaluation, daemon=True)
        thread.start()
    
    def run_evaluation(self):
        try:
            self.queue_log("=" * 60)
            self.queue_log("🚀 EVALUATION STARTED")
            self.queue_log("=" * 60)
            self.queue_log(f"Groups to evaluate: {len(self.groups)}\n")
            
            self.evaluator = UniversalGitHubEvaluator(log_callback=self.queue_log)
            self.all_results = []
            
            for i, (code_url, doc_url) in enumerate(self.groups, 1):
                try:
                    evaluation = self.evaluator.evaluate_group(i, code_url, doc_url)
                    
                    # Créer résultat pour export
                    result = StudentResult(
                        group_number=i,
                        student_name=f"Group {i}",
                        matricule=f"G{i:03d}",
                        code_url=code_url,
                        doc_url=doc_url,
                        note_code=evaluation.note_code,
                        note_documentation=evaluation.note_documentation,
                        note_execution=evaluation.note_execution,
                        note_finale=evaluation.note_finale,
                        code_files_count=len(evaluation.code_files),
                        doc_files_count=len(evaluation.doc_files),
                        compilation_success=any(cf.compilation_success for cf in evaluation.code_files),
                        execution_success=any(cf.execution_result and cf.execution_result.success 
                                             for cf in evaluation.code_files)
                    )
                    
                    self.all_results.append(result)
                    
                    # Sauvegarder JSON
                    self.save_json(evaluation)
                    
                    if i < len(self.groups):
                        self.queue_log("\n⏳ Waiting before next group...\n")
                        time.sleep(2)
                
                except Exception as e:
                    self.queue_log(f"\n❌ Error in group {i}: {str(e)}\n")
                    continue
            
            # Rapport final
            self.queue_log("\n" + "=" * 60)
            self.queue_log("📊 FINAL REPORT")
            self.queue_log("=" * 60)
            self.queue_log(f"✅ Groups evaluated: {len(self.all_results)}/{len(self.groups)}")
            
            if self.all_results:
                avg = sum(r.note_finale for r in self.all_results) / len(self.all_results)
                self.queue_log(f"📈 Average grade: {avg:.2f}/20")
                self.queue_log(f"📊 Best: {max(r.note_finale for r in self.all_results):.2f}/20")
                self.queue_log(f"📉 Worst: {min(r.note_finale for r in self.all_results):.2f}/20")
            
            self.queue_log("\n✅ EVALUATION COMPLETED")
            self.queue_log("=" * 60)
            
            # Activer export
            self.root.after(0, lambda: self.export_btn.config(state='normal'))
            
            self.root.after(0, lambda: messagebox.showinfo(
                "Evaluation Complete",
                f"✅ {len(self.all_results)} group(s) evaluated!\n\n"
                f"📊 Average: {avg:.2f}/20\n\n"
                "Click 'Export Results' to download CSV/Excel"
            ))
        
        except Exception as e:
            self.queue_log(f"\n❌ FATAL ERROR: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
        
        finally:
            if self.evaluator:
                self.evaluator.cleanup()
            
            self.is_evaluating = False
            self.root.after(0, lambda: self.eval_btn.config(state='normal'))
            self.root.after(0, lambda: self.status_label.config(text="● Ready", fg=self.colors['success']))
    
    def save_json(self, evaluation: GroupEvaluation):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"evaluation_group{evaluation.group_number}_{timestamp}.json"
        
        json_data = {
            'metadata': {
                'group': evaluation.group_number,
                'date': datetime.now().isoformat(),
                'code_repo': evaluation.code_repo_url,
                'doc_repo': evaluation.doc_repo_url,
                'analysis_time_seconds': round(evaluation.temps_analyse, 2)
            },
            'grades': {
                'final_grade': round(evaluation.note_finale, 2),
                'code_static': round(evaluation.note_code, 2),
                'code_execution': round(evaluation.note_execution, 2),
                'documentation': round(evaluation.note_documentation, 2),
                'weighting': 'Code 40% + Execution 30% + Documentation 30%'
            },
            'code_files': [
                {
                    'filename': cf.filename,
                    'language': cf.language,
                    'lines_of_code': cf.lines_of_code,
                    'has_comments': cf.has_comments,
                    'has_functions': cf.has_functions,
                    'compilation_success': cf.compilation_success,
                    'execution_success': cf.execution_result.success if cf.execution_result else False,
                    'execution_time': cf.execution_result.execution_time if cf.execution_result else 0,
                    'grade': round(cf.note, 2)
                }
                for cf in evaluation.code_files
            ],
            'documentation_files': [
                {
                    'filename': df.filename,
                    'word_count': df.word_count,
                    'has_structure': df.has_structure,
                    'has_code_blocks': df.has_code_blocks,
                    'technical_content': round(df.technical_content, 2),
                    'grade': round(df.note, 2)
                }
                for df in evaluation.doc_files
            ]
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        self.queue_log(f"💾 Results saved: {filename}")
    
    def export_results(self):
        if not self.all_results:
            messagebox.showwarning("No Results", "No evaluation results to export")
            return
        
        # Dialogue de choix
        export_type = messagebox.askquestion(
            "Export Format",
            "Export to Excel?\n\n"
            "Yes = Excel (.xlsx)\n"
            "No = CSV (.csv)",
            icon='question'
        )
        
        if export_type == 'yes':
            # Export Excel
            if not EXCEL_AVAILABLE:
                messagebox.showerror(
                    "Excel Not Available",
                    "openpyxl is not installed.\n\n"
                    "Install with: pip install openpyxl\n\n"
                    "Exporting as CSV instead."
                )
                self.export_csv()
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"evaluation_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if filename:
                try:
                    ExportManager.export_to_excel(self.all_results, filename)
                    messagebox.showinfo("Success", f"Excel file created:\n{filename}")
                    
                    # Ouvrir le fichier
                    if messagebox.askyesno("Open File", "Open the Excel file now?"):
                        os.startfile(filename) if os.name == 'nt' else os.system(f'open "{filename}"')
                
                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to export:\n{str(e)}")
        else:
            self.export_csv()
    
    def export_csv(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile=f"evaluation_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if filename:
            try:
                ExportManager.export_to_csv(self.all_results, filename)
                messagebox.showinfo("Success", f"CSV file created:\n{filename}")
                
                if messagebox.askyesno("Open File", "Open the CSV file now?"):
                    os.startfile(filename) if os.name == 'nt' else os.system(f'open "{filename}"')
            
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export:\n{str(e)}")

def main():
    root = tk.Tk()
    
    # Centrer la fenêtre
    width = 1200
    height = 750
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    app = MinimalistGitHubEvaluatorGUI(root)
    
    root.mainloop()

if __name__ == "__main__":
    main()